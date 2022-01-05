import openpyxl
wb=openpyxl.load_workbook("Inverter.xlsx")
sheet=wb['Sheet']
wbTemp=openpyxl.Workbook()
sheetTemp=wbTemp['Sheet']
for row in range(1,sheet.max_row+1):
    for col in range(1,sheet.max_column+1):
        #print(row,col)
        sheetTemp.cell(col,row).value=sheet.cell(row,col).value
wbTemp.save("Inverter.xlsx")
print("Done")
