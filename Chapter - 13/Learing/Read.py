import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
file=openpyxl.load_workbook('Book1.xlsx')
sheet=file['Sheet1']
for i in range(1,sheet.max_row):
    for j in range(1,sheet.max_column):
        print("The value cell :",i,get_column_letter(j)," is :-",sheet.cell(i,j).value)
lrow=sheet.max_row
lcol=get_column_letter(sheet.max_column)
lastcel=str(lcol)+str(lrow)
print('\n\n',tuple(sheet['A1':lastcel]),"\n\n")
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value,"\n")
    print('--- END OF ROW ---\n')
