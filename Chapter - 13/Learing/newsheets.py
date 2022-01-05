import openpyxl
wb=openpyxl.Workbook("newsheets.xlsx")
wb.create_sheet(index=1,title="Created sheet")
print(wb.sheetnames)
wb.create_sheet(index=1,title="New Sheeet")
print(wb.sheetnames)
wb.save("newsheets.xlsx")
