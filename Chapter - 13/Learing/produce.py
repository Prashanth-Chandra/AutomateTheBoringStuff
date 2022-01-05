import openpyxl
wb=openpyxl.load_workbook('produceSales.xlsx')
sheet=wb['Sheet']
cnt=0
changes={'Garlic':3.07,'Celery':1.19,'Lemon':1.27}
for rows in range(2,sheet.max_row+1):
    x=sheet.cell(rows,1).value
    if x in changes:
        sheet.cell(rows,2).value=changes[x]
wb.save("updatesProduceSalex.xlsx")
print("Done")
