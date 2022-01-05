import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

wb=openpyxl.Workbook()
sheet=wb['Sheet']
n=int(input("Enter the value of n : "))
boldFont=Font(bold=True)
for i in range(2,n+2):
    sheet.cell(i,1).value=i-1
    x='A'+str(i)
    sheet[x].font=boldFont
    sheet.cell(1,i).value=i-1
    x=str(get_column_letter(i))+str(1)
    sheet[x].font=boldFont
for row in range(2,n+2):
    for col in range(2,n+2):
        sheet.cell(row,col).value=(row-1)*(col-1)
wb.save("multTable.xlsx")
print("Done")
